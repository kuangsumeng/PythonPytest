from Utils.Log_util import log_config_init
from Utils.Global_var import *
import openpyxl as exl
import traceback

log = log_config_init()


class ExcelUtil:

    # Initialize excel file
    def __init__(self, sheet_name, file_name=TEST_CASE_DATA_DIR):
        try:
            self.wb = exl.load_workbook(file_name)
            self.sheet_name = sheet_name
            log.info('Initialize excel file success')
        except Exception:
            log.error('Excel file initialized fail')
            log.error(traceback.format_exc())
            raise

    # Get testcase data of all rows based on sheet name
    def get_sheet_data(self):
        """
        Convert all use cases into dictionary format and add them to the list data_all
        :return: testcase list
        """
        ws = self.wb[self.sheet_name]
        try:
            log.info("Read>>>{}<<<test data success".format(self.sheet_name))
        except Exception:
            log.error("Test data read failure: {}".format(self.sheet_name))
            log.error(traceback.format_exc())
            raise
        else:
            # Read Excel data by traversing the rows and adding each row to a list row_list,
            # ws.rows is equivalent to a generator
            data_list = []
            for row in ws.rows:
                row_list = []
                for cell in row:
                    row_list.append(cell.value)
                data_list.append(row_list)
            # Convert the first line and the non-first line into dictionaries and add them to the list data_all
            data_all = []
            for i in data_list[1:]:
                x = dict(zip(data_list[0], i))
                data_all.append(x)
            self.wb.close()

            return data_all

    def write_excel(self, row, actually, result):
        """
        Write actually and result into the row 、 columns H and I respectively.
        :param row: row need to write in
        :param actually: Test results to be filled in; Ep: pass, failed, blocked
        :param result: Data to writes
        :return:
        """
        ws = self.wb[self.sheet_name]
        ws["I{}".format(row)] = actually
        ws["J{}".format(row)] = result
        self.wb.save(filename=TEST_CASE_DATA_DIR)
        self.wb.close()
